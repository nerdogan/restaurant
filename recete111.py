# -*- coding:utf-8 -*-
__author__ = 'NAMIK'

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Color,Fill
import re
import xlwt
import xlrd
import datetime

def kontrol(girdi):
        girdi = str(girdi)
        ara = re.search("\.", girdi)
        if ara:
            print girdi
            derle = re.compile("\.")
            cikti = derle.sub(",",girdi)
            return float(cikti)
        return int(girdi)




wb =  xlwt.Workbook(encoding="utf-8")
dest_filename = 'empty_book_niss.xls'
date_format = xlwt.XFStyle()
date_format.num_format_str = u'#,##0.00₺'
date_xf = xlwt.easyxf(num_format_str='DD/MM/YYYY')
style1 = xlwt.easyxf('pattern: pattern solid, fore_colour red;')




wb1 = load_workbook('C:\\Users\\NAMIK\\Google Drive\\bishop\\PERSONEL\\ARCHBISHOP.xlsx')
ws = wb1["KAHVALTI"]
aa=0
ab=0
data=[]
#value work properly
print ws['A1'].value #returns foo
print ws['a21'].value #return bar

#formatting does not - THIS IS THE PROBLEM CODE
#print ws.cell('A1').style.font.bold #returns False
#print ws.cell('B1').style.fill.fill_type #returns none
#print ws.cell('B1').style.fill.start_color.index #returns FFFFFFFF

print ws['a1'].has_style #returns true
#but these 2 return the same values! even thought C1 was never set and should be different
print ws['A1'].style


for row in ws.rows:
    ac=0
    ab=ab+1
    if ab==1113:
        break
    if ab < 173:
        continue

    print "   "
    for cell in row:
        if  ( ac==1 or    ac==3):
            if cell.value==None:
                ab=99999
            deger1=cell.value
            data.append(deger1)
            print deger1

            print ab,ac+1
        if ( ac == 2):
            if cell.value == None:
                ab = 99999
            deger1 = float(cell.value)
            data.append(deger1)
            print deger1
            print ab, ac + 1

        if ( ac == 0):
            if cell.value == None:
                ab = 99999
                pass
            deger1 = datetime.datetime.strptime(str(cell.value), "%d.%m.%Y %H:%M")
            print deger1

            t = deger1.timetuple()
            deger2 = datetime.date(t[0],t[1],t[2])
            data.append(deger2)
            print  deger2


        ac=ac+1

elma=0
kesinti="243000000"
print len(data)
satir=3
satir1=1
satir2=1
ws3 =wb.add_sheet("gelir")
ws4 =wb.add_sheet("gider")
for row in range(len(data)/4):



    for col in range(1,2):

        if data[aa + 2] <= 0:
            if kesinti in data[aa+1]:
                elma=elma+data[aa+2]

                aa = aa + 4
                continue
            satir1 = satir1 + 1
            col += 5
            satir = satir1

        else:
            satir2=satir2+1
            satir=satir2
        ws3.write(satir,col, data[aa],date_xf)
        aa=aa+1


        ws3.write(satir,col+1,  data[aa])

        aa=aa+1

        ws3.write(satir,col+2, (data[aa]),date_format)

        aa=aa+1

        ws3.write(satir,col + 3, (data[aa]))

        aa = aa + 1
ws3.write(satir1+1,7,"kredi kart komisyonu")
ws3.write(satir1+1,8,elma)
ws3.write(satir1+4,7,"   Toplam :")
ws3.write(satir1+4,8,xlwt.Formula("sum(I3:I"+str(satir1+3)+")"))
print elma
wb.save(dest_filename)
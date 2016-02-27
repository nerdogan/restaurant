# -*- coding:cp1254 -*-
__author__ = 'NAMIK'

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range


wb = Workbook()
dest_filename = 'empty_book2.xlsx'


wb1 = load_workbook('C:\\Users\\NAMIK\\Google Drive\\bishop\\PERSONEL\\bordrokasım.xlsx', read_only=True)
ws = wb1['Sayfa1']
aa=0
ab=0
data=[]
for row in ws.rows:
    ac=0
    ab=ab+1
    if ab==100000:
        break
    print "   "
    for cell in row:
        if (ab > 11) and (ab%3==0) and (ac==0 or ac==3 or  ac==20):
            if cell.value==None:
                ab=99999
            data.append(cell.value)
            print cell.value

            print ab,ac+1
        ac=ac+1


print len(data)

ws3 = wb.create_sheet(title="Data12")
for row in range(len(data)/3):
    for col in range(1,2):
        ws3.cell(column=col, row=row+1, value="%s" % data[aa])
        aa=aa+1


        ws3.cell(column=col+1, row=row+1, value="%s" %  data[aa])

        aa=aa+1

        ws3.cell(column=col+2, row=row+1, value="%s" %  str(data[aa]))

        aa=aa+1


wb.save(filename = dest_filename)